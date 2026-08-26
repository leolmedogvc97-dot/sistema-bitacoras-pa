import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import json
import os
import random
import pydeck as pdk
from io import BytesIO
from datetime import datetime, date

st.set_page_config(page_title="Sistema Nacional de Bitácoras - Procuraduría Agraria", layout="wide")

# Inyección de estilos CSS personalizados (Colores Institucionales / Guinda Morena)
st.markdown("""
    <style>
    /* Estilos globales y elementos principales */
    h1, h2, h3 {
        color: #6B1D2F !important;
    }
    
    /* Botones principales de la aplicación */
    .stButton>button {
        background-color: #6B1D2F !important;
        color: white !important;
        border-radius: 6px !important;
        border: none !important;
        font-weight: bold;
    }
    
    .stButton>button:hover {
        background-color: #8A253D !important;
        color: white !important;
    }
    
    /* Botones de envío en formularios */
    [data-testid="stFormSubmitButton"]>button {
        background-color: #6B1D2F !important;
        color: white !important;
        border-radius: 6px !important;
        font-weight: bold;
    }
    
    [data-testid="stFormSubmitButton"]>button:hover {
        background-color: #8A253D !important;
    }
    
    /* Barra lateral */
    section[data-testid="stSidebar"] {
        background-color: #faf6f7;
        border-right: 1px solid #e0d0d3;
    }
    
    /* Contenedor de frase agraria */
    .frase-agraria {
        background-color: #fcf5f6;
        border-left: 5px solid #6B1D2F;
        padding: 12px 18px;
        border-radius: 4px;
        margin-bottom: 20px;
        font-style: italic;
        color: #4a2c33;
    }
    </style>
""", unsafe_allow_html=True)

# Archivos persistentes y carpetas de almacenamiento
USUARIOS_FILE = "usuarios.json"
REGISTROS_FILE = "registros.json"
SOLICITUDES_FILE = "solicitudes_gasolina.json"
INCIDENCIAS_FILE = "incidencias_mecanicas.json"
AUDIT_FILE = "audit_log.json"
FOTOS_DIR = "fotos_perfil"
LOGO_FILE = "logo_pa.png"
MUN_FILE = "MUNICIPIOS_202606.xlsx"
LOC_FILE = "LOCALIDADES_202606.xlsx"
os.makedirs(FOTOS_DIR, exist_ok=True)

# Coordenadas de referencia aproximadas para municipios principales
CODS_MUNICIPIOS = {
    "Toluca": {"lat": 19.2826, "lon": -99.6556},
    "Naucalpan de Juárez": {"lat": 19.4781, "lon": -99.2363},
    "Naucalpan": {"lat": 19.4781, "lon": -99.2363},
    "Atizapán de Zaragoza": {"lat": 19.5667, "lon": -99.2500},
    "Tlalnepantla de Baz": {"lat": 19.5400, "lon": -99.1900},
    "Ecatepec de Morelos": {"lat": 19.6010, "lon": -99.0558},
    "Texcoco": {"lat": 19.5100, "lon": -98.8800},
    "Valle de Bravo": {"lat": 19.1944, "lon": -100.1333},
    "Tenancingo": {"lat": 18.9667, "lon": -99.5833},
    "Atlacomulco": {"lat": 19.8000, "lon": -99.8833},
    "Morelia": {"lat": 19.7027, "lon": -101.1924},
    "Uruapan": {"lat": 19.4128, "lon": -102.0647}
}

ESTADOS_REPUBLICA = [
    "Aguascalientes", "Baja California", "Baja California Sur", "Campeche", 
    "Chiapas", "Chihuahua", "Ciudad de México", "Coahuila", "Colima", 
    "Durango", "Estado de México", "Guanajuato", "Guerrero", "Hidalgo", 
    "Jalisco", "Michoacán", "Morelos", "Nayarit", "Nuevo León", "Oaxaca", 
    "Puebla", "Querétaro", "Quintana Roo", "San Luis Potosí", "Sinaloa", 
    "Sonora", "Tabasco", "Tamaulipas", "Tlaxcala", "Veracruz", "Yucatán", "Zacatecas"
]

MUNICIPIOS_EDOMEX = [
    "ACAMBAY DE RUIZ CASTAÑEDA", "ACOLMAN", "ACULCO", "ALMOLOYA DE ALQUISIRAS",
    "ALMOLOYA DE JUÁREZ", "ALMOLOYA DEL RÍO", "AMANALCO", "AMATEPEC", "AMECAMECA",
    "APAXCO", "ATENCO", "ATIZAPÁN", "ATIZAPÁN DE ZARAGOZA", "ATLACOMULCO",
    "ATLAUTLA", "AXAPUSCO", "AYAPANGO", "CALIMAYA", "CAPULHUAC",
    "COACALCO DE BERRIOZÁBAL", "COATEPEC HARINAS", "COCOTITLÁN", "COYOTEPEC",
    "CUAUTITLÁN", "CUAUTITLÁN IZCALLI", "CHALCO", "CHAPA DE MOTA", "CHAPULTEPEC",
    "CHIAUTLA", "CHICOLOAPAN", "CHICONCUAC", "CHIMALHUACÁN", "DONATO GUERRA",
    "ECATEPEC DE MORELOS", "ECATZINGO", "HUEHUETOCA", "HUEYPOXTLA", "HUIXQUILUCAN",
    "ISIDRO FABELA", "IXTAPALUCA", "IXTAPAN DE LA SAL", "IXTAPAN DEL ORO",
    "IXTLAHUACA", "XALATLACO", "JALTENCO", "JILOTEPEC", "JILOTZINGO", "JIQUIPILCO",
    "JOCOTITLÁN", "JOQUICINGO", "JUCHITEPEC", "LERMA", "MALINALCO", "MELCHOR OCAMPO",
    "METEPEC", "MEXICALTZINGO", "MORELOS", "NAUCALPAN DE JUÁREZ", "NEXTLALPAN",
    "NEZAHUALCÓYOTL", "NICOLÁS ROMERO", "NOPALTEPEC", "OCOYOACAC", "OCUILAN",
    "EL ORO", "OTUMBA", "OTZOLOAPAN", "OTZOLOTEPEC", "OZUMBA", "PAPALOTLA",
    "LA PAZ", "POLOTITLÁN", "RAYÓN", "SAN ANTONIO LA ISLA", "SAN FELIPE DEL PROGRESO",
    "SAN MARTÍN DE LAS PIRÁMIDES", "SAN MATEO ATENCO", "SAN SIMÓN DE GUERRERO",
    "SANTO TOMÁS", "SOYANIQUILPAN DE JUÁREZ", "SULTEPEC", "TECÁMAC", "TEJUPILCO",
    "TEMAMATLA", "TEMASCALAPA", "TEMASCALCINGO", "TEMASCALTEPEC", "TEMOAYA",
    "TENANCINGO", "TENANGO DEL AIRE", "TENANGO DEL VALLE", "TEOLOYUCAN",
    "TEOTIHUACÁN", "TEPETLAOXTOC", "TEPETLIXPA", "TEPOTZOTLÁN", "TEQUIXQUIAC",
    "TEXCALTITLÁN", "TEXCALYACAC", "TEXCOCO", "TEZOYUCA", "TIANGUISTENCO",
    "TIMILPAN", "TLALMANALCO", "TLALNEPANTLA DE BAZ", "TLATLAYA", "TOLUCA",
    "TONATICO", "TULTEPEC", "TULTITLÁN", "VALLE DE BRAVO", "VILLA DE ALLENDE",
    "VILLA DEL CARBÓN", "VILLA GUERRERO", "VILLA VICTORIA", "XONACATLÁN",
    "ZACAZONAPAN", "ZACUALPAN", "ZINACANTEPEC", "ZUMPAHUACÁN", "ZUMPANGO",
    "VALLE DE CHALCO SOLIDARIDAD", "LUVIANOS", "SAN JOSÉ DEL RINCÓN", "TONANITLA"
]

JEFATURAS_RESIDENCIA = [
    "RESIDENCIA NAUCALPAN",
    "RESIDENCIA TOLUCA",
    "RESIDENCIA ATLACOMULCO",
    "RESIDENCIA TEXCOCO",
    "RESIDENCIA VALLE DE BRAVO",
    "RESIDENCIA TENANCINGO",
    "RESIDENCIA MORELIA"
]

ROLES_SISTEMA = [
    "Organizador Agrario (Operador)", 
    "Analista de Información", 
    "Administrador de Vehículos (Estatal)",
    "Jefe de Residencia", 
    "Administrador Estatal", 
    "Administrador Nacional"
]

FRASES_AGRARIAS = [
    "“La tierra no pertenece al hombre; el hombre pertenece a la tierra.” — Jefe Seattle",
    "“El agricultor es la única persona en el mundo que gasta dinero en esperar y arriesgarse a cosechar.” — E.W. Howe",
    "“La agricultura es la base de todas las manufacturas y el sustento de la nación.” — Daniel Webster",
    "“Quien planta un árbol cree en el mañana.” — Audrey Hepburn",
    "“El campo da pan a todos, y la justicia agraria da paz y dignidad a nuestra gente.”",
    "“Cultivar la tierra es servir a la patria con las manos y el corazón.”",
    "“La tierra generosa recompensa siempre el esfuerzo honesto de quien la trabaja.”",
    "“El verdadero desarrollo del país florece desde sus raíces ejidales y comunales.”",
    "“Sembrar conciencia en el campo es cosechar soberanía y bienestar social.”",
    "“La labor del agrónomo y del servidor agrario es transformar la esperanza en frutos tangibles.”",
    "“Un pueblo que cuida su campo asegura su porvenir y su libertad.”",
    "“La tierra es el espejo del alma de quienes la trabajan día con día.”",
    "“Servir a los núcleos agrarios es un honor que exige lealtad, vocación y justicia social.”",
    "“Detrás de cada surco hay una historia de esfuerzo, familia y amor por México.”",
    "“La tierra bien administrada y respetada nunca defrauda a sus hijos.”",
    "“Hacer justicia agraria es dar valor institucional al sudor del campesino.”",
    "“El progreso del campo se construye con técnica, compromiso y cercanía con la gente.”",
    "“La naturaleza no hace nada en vano; cada semilla es una promesa de futuro.”",
    "“Trabajar por el campo es sembrar las bases de la justicia y la equidad nacional.”",
    "“Cada kilómetro recorrido en comisión oficial es un paso más hacia la justicia agraria.”",
    "“La tierra es fértil para quien la respeta; la ley es justa para quien la aplica con vocación.”",
    "“El agrarismo es la fuerza viva que alimenta la historia y el orgullo de nuestras comunidades.”",
    "“Cultivar el campo con ciencia y conciencia es el mejor legado para las nuevas generaciones.”",
    "“El servicio público en el sector agrario es vocación de entrega y transformación social.”",
    "“Donde hay un ejidatario trabajando, hay una patria entera floreciendo.”"
]

@st.cache_data
def cargar_catalogos_geograficos():
    muns_df = pd.read_excel(MUN_FILE) if os.path.exists(MUN_FILE) else pd.DataFrame()
    locs_df = pd.read_excel(LOC_FILE) if os.path.exists(LOC_FILE) else pd.DataFrame()
    return muns_df, locs_df

muns_global, locs_global = cargar_catalogos_geograficos()

def obtener_municipios_estado(estado_nombre):
    if muns_global.empty or estado_nombre not in ESTADOS_REPUBLICA:
        return []
    efe_key = ESTADOS_REPUBLICA.index(estado_nombre) + 1
    if 'EFE_KEY' in muns_global.columns:
        muns = muns_global[muns_global['EFE_KEY'] == efe_key]['MUNICIPIO'].dropna().tolist()
    elif 'ENTIDAD' in muns_global.columns:
        muns = muns_global[muns_global['ENTIDAD'].str.upper() == estado_nombre.upper()]['MUNICIPIO'].dropna().tolist()
    else:
        muns = muns_global.iloc[:, 1].dropna().tolist()
    return sorted(list(set(muns)))

def obtener_localidades_municipio(estado_nombre, municipio_nombre):
    if locs_global.empty or estado_nombre not in ESTADOS_REPUBLICA:
        return ["Cabecera Municipal"]
    efe_key = ESTADOS_REPUBLICA.index(estado_nombre) + 1
    
    df_locs = locs_global.copy()
    if 'EFE_KEY' in df_locs.columns and 'MUNICIPIO' in df_locs.columns:
        locs = df_locs[(df_locs['EFE_KEY'] == efe_key) & (df_locs['MUNICIPIO'].str.upper() == municipio_nombre.upper())]['LOCALIDAD'].dropna().tolist()
    elif 'MUNICIPIO' in df_locs.columns:
        locs = df_locs[df_locs['MUNICIPIO'].str.upper() == municipio_nombre.upper()]['LOCALIDAD'].dropna().tolist()
    else:
        locs = df_locs.iloc[:, 2].dropna().tolist()
        
    if not locs:
        return ["Cabecera Municipal"]
    return sorted(list(set(locs)))

def cargar_usuarios():
    usuarios_base = {
        "victor.olmedo@pa.gob.mx": {"nombre": "VÍCTOR LEONARDO OLMEDO GONZALEZ", "pass": "Leonardo", "licencia": "0101P3402484l", "rol": "Administrador Nacional", "estado": "Estado de México", "jefatura": "RESIDENCIA NAUCALPAN", "jefe_residencia": "N/A", "foto": "", "activo": True},
        "marichuy.duarte@pa.gob.mx": {"nombre": "MARICHUY DUARTE SALAMANCA", "pass": "Marichuy2026", "licencia": "12345678", "rol": "Administrador Nacional", "estado": "Estado de México", "jefatura": "RESIDENCIA TOLUCA", "jefe_residencia": "N/A", "foto": "", "activo": True},
        "marichuy@pa.gob.mx": {"nombre": "MARICHUY", "pass": "Marichuy2026", "licencia": "0000000000000", "rol": "Administrador Nacional", "estado": "Michoacán", "jefatura": "RESIDENCIA MORELIA", "jefe_residencia": "N/A", "foto": "", "activo": True},
        "jcarlos.patino@pa.gob.mx": {"nombre": "JUAN CARLOS PATINO PEREZ", "pass": "Perez", "licencia": "", "rol": "Administrador Nacional", "estado": "Estado de México", "jefatura": "RESIDENCIA TOLUCA", "jefe_residencia": "NINGUNO", "foto": "", "activo": True},
        "arosales@pa.gob.mx": {"nombre": "ARISVE LESEIE ESLAVA ROSALES", "pass": "Arosales", "licencia": "", "rol": "Administrador Nacional", "estado": "Estado de México", "jefatura": "RESIDENCIA TOLUCA", "jefe_residencia": "NINGUNO", "foto": "", "activo": True},
        "carmen.lara@pa.gob.mx": {"nombre": "CARMEN LARA", "pass": "Carmen", "licencia": "", "rol": "Administrador Nacional", "estado": "Aguascalientes", "jefatura": "RESIDENCIA NAUCALPAN", "jefe_residencia": "ING. GABRIEL ESTRADA", "foto": "", "activo": True},
        "josue.rodriguez@pa.gob.mx": {"nombre": "JOSUE RODRIGUEZ", "pass": "Josue", "licencia": "", "rol": "Administrador Nacional", "estado": "Aguascalientes", "jefatura": "RESIDENCIA NAUCALPAN", "jefe_residencia": "ING. GABRIEL ESTRADA", "foto": "", "activo": True},
        "nsalgado@pa.gob.mx": {"nombre": "NANCY SALGADO ANTUNEZ", "pass": "Nancy", "licencia": "", "rol": "Analista de Información", "estado": "Estado de México", "jefatura": "RESIDENCIA NAUCALPAN", "jefe_residencia": "ING. GABRIEL ESTRADA", "foto": "", "activo": True},
        "esperanza.ramos@pa.gob.mx": {"nombre": "ESPERANZA WENDY RAMOS RODRIGUEZ", "pass": "Wendy", "licencia": "", "rol": "Administrador Nacional", "estado": "Estado de México", "jefatura": "RESIDENCIA NAUCALPAN", "jefe_residencia": "ADAN JIMENEZ", "foto": "", "activo": True},
        "carlos.javg.96@gmail.com": {"nombre": "CARLOS JAVIER GALVEZ GONZALEZ", "pass": "Carlos", "licencia": "", "rol": "Administrador Nacional", "estado": "Estado de México", "jefatura": "RESIDENCIA NAUCALPAN", "jefe_residencia": "ADAN JIMENEZ", "foto": "", "activo": True},
        "dehnny.vasquez@pa.gob.mx": {"nombre": "DEHNNY VAZQUEZ FLORES", "pass": "Dehnny", "licencia": "", "rol": "Organizador Agrario (Operador)", "estado": "Estado de México", "jefatura": "RESIDENCIA TOLUCA", "jefe_residencia": "", "foto": "", "activo": True}
    }
    
    if os.path.exists(USUARIOS_FILE):
        try:
            with open(USUARIOS_FILE, "r", encoding="utf-8") as f:
                usuarios_guardados = json.load(f)
                for email, data in usuarios_base.items():
                    if email not in usuarios_guardados:
                        usuarios_guardados[email] = data
                return usuarios_guardados
        except:
            pass
    return usuarios_base

def guardar_usuarios(usuarios_dict):
    with open(USUARIOS_FILE, "w", encoding="utf-8") as f:
        json.dump(usuarios_dict, f, ensure_ascii=False, indent=4)

def cargar_registros_acumulados():
    if os.path.exists(REGISTROS_FILE):
        try:
            with open(REGISTROS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return []

def guardar_registros_acumulados(registros_list):
    with open(REGISTROS_FILE, "w", encoding="utf-8") as f:
        json.dump(registros_list, f, ensure_ascii=False, indent=4)

def cargar_solicitudes():
    if os.path.exists(SOLICITUDES_FILE):
        try:
            with open(SOLICITUDES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return []

def guardar_solicitudes(solicitudes_list):
    with open(SOLICITUDES_FILE, "w", encoding="utf-8") as f:
        json.dump(solicitudes_list, f, ensure_ascii=False, indent=4)

def cargar_incidencias():
    if os.path.exists(INCIDENCIAS_FILE):
        try:
            with open(INCIDENCIAS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return []

def guardar_incidencias(incidencias_list):
    with open(INCIDENCIAS_FILE, "w", encoding="utf-8") as f:
        json.dump(incidencias_list, f, ensure_ascii=False, indent=4)

def registrar_auditoria(accion, detalle):
    log_entry = {
        "FECHA_HORA": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "USUARIO": st.session_state.get("current_email", "Sistema"),
        "ROL": st.session_state.get("current_rol", "N/A"),
        "ACCION": accion,
        "DETALLE": detalle
    }
    logs = []
    if os.path.exists(AUDIT_FILE):
        try:
            with open(AUDIT_FILE, "r", encoding="utf-8") as f:
                logs = json.load(f)
        except:
            pass
    logs.append(log_entry)
    with open(AUDIT_FILE, "w", encoding="utf-8") as f:
        json.dump(logs, f, ensure_ascii=False, indent=4)

OPCIONES_GASOLINA = {
    "🔴 1/4 de Tanque": "1/4",
    "🟡 1/2 Tanque": "1/2",
    "🟢 3/4 de Tanque": "3/4",
    "🟢 Tanque Lleno (1/1)": "1/1",
    "🔴 Reserva / Vacío (V)": "V"
}

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
if "current_email" not in st.session_state:
    st.session_state["current_email"] = ""
if "current_user" not in st.session_state:
    st.session_state["current_user"] = ""
if "current_licencia" not in st.session_state:
    st.session_state["current_licencia"] = ""
if "current_rol" not in st.session_state:
    st.session_state["current_rol"] = ""
if "current_estado" not in st.session_state:
    st.session_state["current_estado"] = ""
if "current_jefatura" not in st.session_state:
    st.session_state["current_jefatura"] = ""
if "current_jefe_residencia" not in st.session_state:
    st.session_state["current_jefe_residencia"] = ""

if "frase_dia" not in st.session_state:
    st.session_state["frase_dia"] = random.choice(FRASES_AGRARIAS)

# --- PANTALLA DE LOGIN ---
if not st.session_state["logged_in"]:
    col_l_logo, col_l_title = st.columns([0.15, 2.85])
    with col_l_logo:
        if os.path.exists(LOGO_FILE):
            st.image(LOGO_FILE, width=60)
    with col_l_title:
        st.title("Acceso al Sistema Nacional de Bitácoras")
    st.markdown("---")
    
    col_l1, col_l2, col_l3 = st.columns([1, 2, 1])
    with col_l2:
        st.subheader("Iniciar Sesión")
        email_input = st.text_input("Correo Electrónico Institucional", max_chars=300).strip().lower()
        pass_input = st.text_input("Contraseña", type="password", max_chars=300)
        
        if st.button("🔑 Ingresar al Sistema", use_container_width=True):
            usuarios_actuales = cargar_usuarios()
            if email_input in usuarios_actuales and usuarios_actuales[email_input]["pass"] == pass_input:
                if not usuarios_actuales[email_input].get("activo", True):
                    st.error("⚠️ Tu cuenta se encuentra desactivada. Contacta al Administrador Nacional.")
                else:
                    st.session_state["logged_in"] = True
                    st.session_state["current_email"] = email_input
                    st.session_state["current_user"] = usuarios_actuales[email_input]["nombre"]
                    st.session_state["current_licencia"] = usuarios_actuales[email_input]["licencia"]
                    st.session_state["current_rol"] = usuarios_actuales[email_input].get("rol", "Organizador Agrario (Operador)")
                    st.session_state["current_estado"] = usuarios_actuales[email_input].get("estado", "Estado de México")
                    st.session_state["current_jefatura"] = usuarios_actuales[email_input].get("jefatura", "RESIDENCIA NAUCALPAN")
                    st.session_state["current_jefe_residencia"] = usuarios_actuales[email_input].get("jefe_residencia", "N/A")
                    st.session_state["frase_dia"] = random.choice(FRASES_AGRARIAS)
                    registrar_auditoria("INICIO DE SESION", f"Acceso exitoso de {email_input}")
                    st.success("¡Acceso concedido! Cargando sistema...")
                    st.rerun()
            else:
                st.error("⚠️ Usuario o contraseña incorrectos. Verifica tus datos.")
    st.stop()

# --- APLICACIÓN PRINCIPAL ---
col_m_logo, col_m_title = st.columns([0.15, 2.85])
with col_m_logo:
    if os.path.exists(LOGO_FILE):
        st.image(LOGO_FILE, width=60)
with col_m_title:
    st.title("Sistema Nacional de Control Vehicular y Bitácoras")

st.markdown(f'<div class="frase-agraria">🌾 <b>Reflexión Agraria del Día:</b> {st.session_state["frase_dia"]}</div>', unsafe_allow_html=True)
st.markdown("---")

if os.path.exists(LOGO_FILE):
    st.sidebar.image(LOGO_FILE, use_container_width=True)

current_email_key = st.session_state['current_email']
usuarios_actuales_sidebar = cargar_usuarios()

rol_actual = st.session_state.get("current_rol", "Organizador Agrario (Operador)")

st.sidebar.title("📌 Menú de Navegación")
if rol_actual == "Administrador Nacional":
    modulos_disponibles = ["Módulo de Captura (Recorrido)", "Mi Perfil / Foto", "Solicitud de Recurso de Gasolina", "Reporte de Incidencias en Ruta", "Panel de Administración y Auditoría"]
elif rol_actual in ["Administrador Estatal", "Administrador de Vehículos (Estatal)", "Jefe de Residencia", "Analista de Información"]:
    modulos_disponibles = ["Módulo de Captura (Recorrido)", "Mi Perfil / Foto", "Solicitud de Recurso de Gasolina", "Reporte de Incidencias en Ruta", "Panel de Supervisión (Estatal/Residencia)"]
else:
    modulos_disponibles = ["Módulo de Captura (Recorrido)", "Mi Perfil / Foto", "Solicitud de Recurso de Gasolina", "Reporte de Incidencias en Ruta"]

perfil = st.sidebar.radio("Selecciona tu módulo:", modulos_disponibles)

st.sidebar.markdown("---")
with st.sidebar.expander("👤 Sesión Activa (Ver Datos)", expanded=False):
    foto_actual = usuarios_actuales_sidebar.get(current_email_key, {}).get("foto", "")
    if foto_actual and os.path.exists(foto_actual):
        st.image(foto_actual, width=100)
    else:
        st.info("Sin foto de perfil.")

    st.write(f"**Usuario:** {st.session_state['current_user']}")
    st.write(f"**Correo:** {st.session_state['current_email']}")
    st.write(f"**Rol:** {st.session_state.get('current_rol', 'Organizador')}")
    st.write(f"**Estado:** {st.session_state.get('current_estado', 'N/A')}")
    st.write(f"**Jefatura:** {st.session_state.get('current_jefatura', 'N/A')}")
    st.write(f"**Jefe Resp.:** {st.session_state.get('current_jefe_residencia', 'N/A')}")

st.sidebar.markdown("---")
if st.sidebar.button("🚪 Cerrar Sesión"):
    registrar_auditoria("CIERRE DE SESION", f"Cierre de sesión de {current_email_key}")
    st.session_state["logged_in"] = False
    st.session_state["current_email"] = ""
    st.rerun()

if perfil == "Mi Perfil / Foto":
    st.subheader("🖼️ Configuración de Perfil y Adscripción Institucional")
    st.markdown("Actualiza tu fotografía de perfil y verifica tu pertenencia territorial, jefatura y jefe de residencia correspondiente.")
    
    usuarios_dict_perfil = cargar_usuarios()
    datos_u_actual = usuarios_dict_perfil.get(current_email_key, {})
    
    st.info(f"""
    * **Estado de Adscripción:** {datos_u_actual.get('estado', 'N/A')}
    * **Jefatura de Residencia:** {datos_u_actual.get('jefatura', 'N/A')}
    * **Jefe de Residencia Asignado:** {datos_u_actual.get('jefe_residencia', 'N/A')}
    * **Rol en la Red:** {datos_u_actual.get('rol', 'N/A')}
    """)
    
    with st.form("form_foto_perfil"):
        archivo_foto = st.file_uploader("Seleccionar imagen de perfil (JPG, PNG)", type=["jpg", "jpeg", "png"])
        btn_subir_foto = st.form_submit_button("💾 Guardar Fotografía")
        
        if btn_subir_foto:
            if archivo_foto is not None:
                extension = archivo_foto.name.split(".")[-1]
                nombre_archivo = f"{current_email_key.replace('@', '_').replace('.', '_')}.{extension}"
                ruta_destino = os.path.join(FOTOS_DIR, nombre_archivo)
                
                with open(ruta_destino, "wb") as f:
                    f.write(archivo_foto.getbuffer())
                
                all_users = cargar_usuarios()
                if current_email_key in all_users:
                    all_users[current_email_key]["foto"] = ruta_destino
                    guardar_usuarios(all_users)
                    registrar_auditoria("ACTUALIZACION FOTO", f"Actualización de foto para {current_email_key}")
                    st.success("¡Fotografía de perfil actualizada con éxito!")
            else:
                st.warning("⚠️ Selecciona un archivo de imagen válido antes de guardar.")

elif perfil == "Solicitud de Recurso de Gasolina":
    st.subheader("⛽ Solicitud de Recurso de Gasolina para Comisión Oficial")
    st.markdown("Apartado institucional para la gestión, solicitud y aprobación de asignación presupuestal de combustible.")
    
    todas_sols_notif = cargar_solicitudes()
    sols_mias = [s for s in todas_sols_notif if s.get("CORREO") == current_email_key]
    for s in sols_mias:
        est = s.get("ESTATUS", "")
        if est == "APROBADO":
            st.success(f"🎉 ¡Notificación! Tu solicitud para el destino **{s.get('DESTINO')}** ha sido **APROBADA**.")
        elif est == "RECHAZADO":
            st.error(f"⚠️ Aviso: Tu solicitud para el destino **{s.get('DESTINO')}** fue **RECHAZADA**.")

    estado_usuario_actual = st.session_state.get("current_estado", "Estado de México")
    jefatura_actual = st.session_state.get("current_jefatura", "RESIDENCIA NAUCALPAN")
    solicitante_actual = st.session_state.get("current_user", "")
    
    lista_municipios = obtener_municipios_estado(estado_usuario_actual)
    if not lista_municipios:
        if estado_usuario_actual == "Estado de México":
            lista_municipios = MUNICIPIOS_EDOMEX
        else:
            lista_municipios = ["Cabecera Municipal"]
        
    with st.form("form_solicitud_gasolina"):
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            fecha_solicitud = st.date_input("Fecha de Solicitud", value=date.today())
            analista_solicitante = st.text_input("Analista de Información / Solicitante", value=solicitante_actual, max_chars=300)
            residencia_adscripcion = st.text_input("Jefatura de Residencia", value=jefatura_actual, max_chars=300)
            funcionario_comisionado = st.text_input("Funcionario / Organizador Asignado a Comisión", value="", max_chars=300)
        with col_s2:
            municipio_destino = st.selectbox(f"Municipio de Destino ({estado_usuario_actual})", lista_municipios)
            
            lista_loc_com = obtener_localidades_municipio(estado_usuario_actual, municipio_destino)
            if not lista_loc_com:
                lista_loc_com = ["Cabecera Municipal"]
            localidad_destino = st.selectbox("Localidad / Poblado de Destino", lista_loc_com)
            
            vehiculo_asignado = st.selectbox("Vehículo Oficial Asignado", ["NISSAN VERSA", "PickUp", "Estacas"])
            placas_vehiculo = st.text_input("Placas del Vehículo", value="MGX-543-A", max_chars=300)
            
        st.markdown("---")
        col_s3, col_s4, col_s5 = st.columns(3)
        with col_s3:
            f_inicio_com = st.date_input("Fecha Inicio de Comisión", value=date.today())
        with col_s4:
            f_fin_com = st.date_input("Fecha Término de Comisión", value=date.today())
        with col_s5:
            monto_solicitado = st.number_input("Monto Solicitado para Combustible ($ MXN)", min_value=0.0, value=1500.0, step=50.0)
            
        col_s6, col_s7 = st.columns(2)
        with col_s6:
            oficio_asociado = st.text_input("Número de Oficio de Comisión", value="", max_chars=300)
        with col_s7:
            motivo_comision = st.text_input("Motivo / Descripción de la Comisión Oficial", value="", max_chars=300)
            
        btn_enviar_solicitud = st.form_submit_button("📥 Enviar Solicitud de Recurso Estatal")
        
        if btn_enviar_solicitud:
            if not funcionario_comisionado.strip() or not motivo_comision.strip():
                st.error("⚠️ Debes completar el nombre del funcionario comisionado y el motivo.")
            else:
                nueva_solicitud = {
                    "FECHA SOLICITUD": fecha_solicitud.strftime("%d/%m/%Y"),
                    "SOLICITANTE": analista_solicitante,
                    "ROL SOLICITANTE": rol_actual,
                    "JEFATURA": residencia_adscripcion,
                    "ESTADO": estado_usuario_actual,
                    "FUNCIONARIO": funcionario_comisionado.upper(),
                    "DESTINO": f"{localidad_destino}, {municipio_destino}",
                    "VEHICULO": vehiculo_asignado,
                    "PLACAS": placas_vehiculo,
                    "FECHA INICIO": f_inicio_com.strftime("%d/%m/%Y"),
                    "FECHA TERMINO": f_fin_com.strftime("%d/%m/%Y"),
                    "MONTO SOLICITADO": monto_solicitado,
                    "OFICIO": oficio_asociado if oficio_asociado else "N/A",
                    "MOTIVO": motivo_comision,
                    "ESTATUS": "PENDIENTE DE APROBACIÓN ESTATAL",
                    "CORREO": current_email_key
                }
                
                lista_sols = cargar_solicitudes()
                lista_sols.append(nueva_solicitud)
                guardar_solicitudes(lista_sols)
                registrar_auditoria("SOLICITUD GASOLINA", f"Solicitud de ${monto_solicitado} para {funcionario_comisionado} en {estado_usuario_actual}")
                st.success("✅ Solicitud de recurso de gasolina enviada y registrada correctamente para el estado.")

    solicitudes_historial = cargar_solicitudes()
    if len(solicitudes_historial) > 0:
        st.markdown("---")
        st.subheader("📋 Historial y Gestión de Solicitudes de Gasolina")
        df_sols = pd.DataFrame(solicitudes_historial)
        
        if rol_actual == "Analista de Información":
            df_sols_filtrado = df_sols[df_sols['ESTADO'] == estado_usuario_actual]
        elif rol_actual in ["Jefe de Residencia", "Administrador de Vehículos (Estatal)"]:
            df_sols_filtrado = df_sols[df_sols['JEFATURA'] == jefatura_actual]
        elif rol_actual in ["Administrador Estatal", "Administrador Nacional"]:
            df_sols_filtrado = df_sols
        else:
            df_sols_filtrado = df_sols[df_sols['CORREO'] == current_email_key]
            
        st.dataframe(df_sols_filtrado, use_container_width=True)
        
        if rol_actual in ["Jefe de Residencia", "Administrador de Vehículos (Estatal)", "Administrador Estatal", "Administrador Nacional"]:
            st.markdown("---")
            st.subheader("⚙️ Panel de Aprobación de Solicitudes")
            for idx, sol in enumerate(solicitudes_historial):
                if sol.get("ESTATUS") == "PENDIENTE DE APROBACIÓN ESTATAL":
                    col_info_s, col_btn_ap, col_btn_re = st.columns([5, 1, 1])
                    with col_info_s:
                        st.write(f"**{sol['FUNCIONARIO']}** | Destino: {sol['DESTINO']} | Monto: ${sol['MONTO SOLICITADO']:,.2f} MXN ({sol['JEFATURA']})")
                    with col_btn_ap:
                        if st.button("✅ Aprobar", key=f"aprobar_{idx}"):
                            solicitudes_historial[idx]["ESTATUS"] = "APROBADO"
                            guardar_solicitudes(solicitudes_historial)
                            registrar_auditoria("APROBAR SOLICITUD", f"Aprobación de recurso para {sol['FUNCIONARIO']}")
                            st.rerun()
                    with col_btn_re:
                        if st.button("❌ Rechazar", key=f"rechazar_{idx}"):
                            solicitudes_historial[idx]["ESTATUS"] = "RECHAZADO"
                            guardar_solicitudes(solicitudes_historial)
                            registrar_auditoria("RECHAZAR SOLICITUD", f"Rechazo de recurso para {sol['FUNCIONARIO']}")
                            st.rerun()

elif perfil == "Reporte de Incidencias en Ruta":
    st.subheader("🚨 Reporte y Gestión de Incidencias y Fallas Mecánicas")
    st.markdown("Módulo para registrar y dar seguimiento a averías, fallas mecánicas o imprevistos durante las comisiones oficiales.")
    
    estado_usuario_actual = st.session_state.get("current_estado", "Estado de México")
    jefatura_actual = st.session_state.get("current_jefatura", "RESIDENCIA NAUCALPAN")
    usuario_actual_nombre = st.session_state.get("current_user", "")
    
    with st.form("form_incidencia"):
        col_i1, col_i2 = st.columns(2)
        with col_i1:
            fecha_inc = st.date_input("Fecha de la Incidencia", value=date.today())
            placas_inc = st.text_input("Placas del Vehículo Afectado", value="MGX-543-A", max_chars=300)
            tipo_falla = st.selectbox("Tipo de Incidencia / Falla", [
                "🔴 Falla Mecánica Mayor (Motor/Transmisión)",
                "🟡 Falla Eléctrica / Batería",
                "🟠 Pinchadura de Neumático / Llanta",
                "🔵 Calentamiento / Sistema de Enfriamiento",
                "⚪ Otro imprevisto en ruta"
            ])
        with col_i2:
            municipio_inc = st.text_input("Municipio / Lugar de la Avería", value="", max_chars=300)
            gravedad_inc = st.selectbox("Nivel de Urgencia", ["Baja", "Media", "Alta (Requiere Grúa/Auxilio Inmediato)"])
            
        descripcion_inc = st.text_area("Descripción detallada de la incidencia mecánica o imprevisto")
        
        btn_enviar_inc = st.form_submit_button("🚨 Enviar Reporte de Incidencia")
        if btn_enviar_inc:
            if not municipio_inc.strip() or not descripcion_inc.strip():
                st.error("⚠️ Debes completar el municipio y la descripción de la falla.")
            else:
                nueva_inc = {
                    "FECHA": fecha_inc.strftime("%d/%m/%Y"),
                    "CONDUCTOR": usuario_actual_nombre,
                    "CORREO": current_email_key,
                    "JEFATURA": jefatura_actual,
                    "ESTADO": estado_usuario_actual,
                    "PLACAS": placas_inc.upper(),
                    "TIPO": tipo_falla,
                    "MUNICIPIO": municipio_inc.upper(),
                    "URGENCIA": gravedad_inc,
                    "DESCRIPCION": descripcion_inc,
                    "ESTATUS": "PENDIENTE DE ATENCIÓN"
                }
                lista_incs = cargar_incidencias()
                lista_incs.append(nueva_inc)
                guardar_incidencias(lista_incs)
                registrar_auditoria("REPORTE INCIDENCIA", f"Falla reportada en vehículo {placas_inc} ({tipo_falla})")
                st.success("✅ Incidencia reportada exitosamente. Se ha notificado al Administrador de Vehículos y Jefatura.")

    todas_incs = cargar_incidencias()
    if todas_incs:
        st.markdown("---")
        st.subheader("📋 Historial de Incidencias en Ruta")
        df_incs = pd.DataFrame(todas_incs)
        
        if rol_actual in ["Jefe de Residencia", "Administrador de Vehículos (Estatal)"]:
            df_incs_filtrado = df_incs[df_incs['ESTADO'] == estado_usuario_actual]
        elif rol_actual == "Administrador Estatal":
            df_incs_filtrado = df_incs[df_incs['ESTADO'] == estado_usuario_actual]
        elif rol_actual == "Administrador Nacional":
            df_incs_filtrado = df_incs
        else:
            df_incs_filtrado = df_incs[df_incs['CORREO'] == current_email_key]
            
        st.dataframe(df_incs_filtrado, use_container_width=True)
        
        if rol_actual in ["Administrador de Vehículos (Estatal)", "Administrador Estatal", "Administrador Nacional", "Jefe de Residencia"]:
            st.markdown("---")
            st.subheader("⚙️ Panel de Gestión de Mantenimiento Correctivo")
            for idx, inc in enumerate(todas_incs):
                if inc.get("ESTATUS") == "PENDIENTE DE ATENCIÓN":
                    col_inf_inc, col_btn_res = st.columns([5, 2])
                    with col_inf_inc:
                        st.warning(f"🚨 **Placas:** {inc['PLACAS']} | **Tipo:** {inc['TIPO']} | **Estado/Lugar:** {inc['ESTADO']} - {inc['MUNICIPIO']} | **Urgencia:** {inc['URGENCIA']}\n\n*{inc['DESCRIPCION']}*")
                    with col_btn_res:
                        if st.button("🛠️ Marcar Atendido / Resolver", key=f"resolver_inc_{idx}"):
                            todas_incs[idx]["ESTATUS"] = "RESUELTO / MANTENIMIENTO AUTORIZADO"
                            guardar_incidencias(todas_incs)
                            registrar_auditoria("RESOLVER INCIDENCIA", f"Incidencia resuelta para vehículo {inc['PLACAS']}")
                            st.rerun()

elif perfil == "Módulo de Captura (Recorrido)":
    st.subheader("📝 Módulo de Captura por Día - Organizador / Operador")
    estado_usuario_actual = st.session_state.get("current_estado", "Estado de México")
    jefatura_actual = st.session_state.get("current_jefatura", "RESIDENCIA NAUCALPAN")
    jefe_actual = st.session_state.get("current_jefe_residencia", "N/A")
    
    st.markdown(f"Ingresa los datos de tu recorrido. Ubicación filtrada para **{estado_usuario_actual}** | Jefatura: **{jefatura_actual}** | Jefe: **{jefe_actual}**.")
    
    registros_previos_user = [r for r in cargar_registros_acumulados() if r.get("CORREO_ORGANIZADOR") == current_email_key]
    km_sugerido = 0.0
    if registros_previos_user:
        ultimo_reg = registros_previos_user[-1]
        km_sugerido = float(ultimo_reg.get("KM FINAL / Km de Llegada", 0.0))

    lista_municipios = obtener_municipios_estado(estado_usuario_actual)
    if not lista_municipios:
        if estado_usuario_actual == "Estado de México":
            lista_municipios = MUNICIPIOS_EDOMEX
        else:
            lista_municipios = ["Cabecera Municipal"]
        
    with st.form("form_captura_dia"):
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            fecha = st.date_input("Fecha de registro del uso del vehículo")
            municipio = st.selectbox(f"Municipio ({estado_usuario_actual})", lista_municipios)
            
            lista_localidades = obtener_localidades_municipio(estado_usuario_actual, municipio)
            if not lista_localidades:
                lista_localidades = ["Sin localidades registradas en catálogo"]
            poblado = st.selectbox("Poblado / Localidad", lista_localidades)
            
            folio_ciia = st.text_input("Folio CIIA", value="", max_chars=300)
        with col2:
            h_salida = st.text_input("Hora de Salida (Formato 24h, ej. 09:00)", value="09:00", max_chars=300)
            km_inicial = st.number_input("KM Inicial / Salida", min_value=0.0, value=km_sugerido, step=1.0)
            h_llegada = st.text_input("Hora de Llegada (Formato 24h, ej. 17:00)", value="17:00", max_chars=300)
            km_final = st.number_input("KM Final / Llegada", min_value=0.0, value=km_sugerido, step=1.0)
        with col3:
            lista_adscripciones_unicas = sorted(list(set([
                jefatura_actual, 
                "RESIDENCIA NAUCALPAN", 
                "RESIDENCIA TOLUCA", 
                "RESIDENCIA ATLACOMULCO", 
                "RESIDENCIA TEXCOCO", 
                "RESIDENCIA VALLE DE BRAVO", 
                "RESIDENCIA TENANCINGO"
            ])))
            residencia = st.selectbox("Área de Adscripción", lista_adscripciones_unicas)
            vehiculo = st.selectbox("Tipo de Vehículo", ["NISSAN VERSA", "PickUp", "Estacas"])
            placas = st.text_input("Placas", value="MGX-543-A", max_chars=300)
            licencia = st.text_input("No. De Licencia", value=st.session_state["current_licencia"], max_chars=300)
        with col4:
            usuario = st.text_input("Usuario Responsable", value=st.session_state["current_user"], max_chars=300)
            dotacion = st.number_input("Dotación de Gasolina ($)", min_value=0.0, value=200.0, step=1.0)
            
            gas_salida_label = st.selectbox("Gasolina de Salida (Nivel)", list(OPCIONES_GASOLINA.keys()))
            gas_llegada_label = st.selectbox("Gasolina de Llegada (Nivel)", list(OPCIONES_GASOLINA.keys()))
        
        st.markdown("---")
        col_o1, col_o2, col_o3 = st.columns(3)
        with col_o1:
            oficio_num = st.text_input("Oficio Número (Opcional)", value="", max_chars=300)
        with col_o2:
            oficio_fecha = st.text_input("Oficio Fecha (Opcional)", value="", max_chars=300)
        with col_o3:
            observaciones = st.text_input("Observaciones / Ruta", value="", max_chars=300)
        
        guardar_dia = st.form_submit_button("💾 Guardar Día")
        
        if guardar_dia:
            if h_salida.strip() == h_llegada.strip():
                st.error("⚠️ Error: La Hora de Salida y la Hora de Llegada no pueden ser iguales en formato 24 horas.")
            elif km_final < km_inicial:
                st.warning("⚠️ Aviso: El KM Final no puede ser menor al KM Inicial.")
            else:
                fecha_str = fecha.strftime("%d/%m/%Y")
                placas_upper = placas.strip().upper()
                registros_actuales = cargar_registros_acumulados()
                
                conflicto_duplicado = False
                mensaje_error = ""
                for r in registros_actuales:
                    r_fecha = r.get("FECHA COMPLETA")
                    r_placas = str(r.get("Placas", "")).strip().upper()
                    r_km_ini = float(r.get("KM INICIAL / Km de Salida", 0))
                    r_km_fin = float(r.get("KM FINAL / Km de Llegada", 0))
                    
                    if r_fecha == fecha_str and r_placas == placas_upper:
                        if r_km_ini == km_inicial and r_km_fin == km_final and str(r.get("MUNICIPIO")) == municipio:
                            conflicto_duplicado = True
                            mensaje_error = "⚠️ Registro duplicado: Ya existe un recorrido guardado con exactamente la misma información (misma fecha, mismas placas, mismo rango de kilometraje y municipio)."
                            break
                        elif km_inicial < r_km_fin and km_final > r_km_ini:
                            conflicto_duplicado = True
                            mensaje_error = f"⚠️ Conflicto de Kilometraje: El vehículo con placas {placas_upper} ya registra un recorrido en la fecha {fecha_str} que abarca de {r_km_ini:,.0f} a {r_km_fin:,.0f} km. El rango ingresado se traslapa."
                            break

                if conflicto_duplicado:
                    st.error(mensaje_error)
                else:
                    recorrido = km_final - km_inicial
                    nuevo_reg = {
                        "FECHA COMPLETA": fecha_str,
                        "MES": fecha.strftime("%B").upper(),
                        "MUNICIPIO": municipio,
                        "POBLADO": poblado,
                        "folio CIIA": folio_ciia,
                        "HORA DE SALIDA": h_salida,
                        "KM INICIAL / Km de Salida": km_inicial,
                        "RECORRIDO": recorrido,
                        "HORA DE LLEGADA": h_llegada,
                        "KM FINAL / Km de Llegada": km_final,
                        "GASTO COMBUSTI": dotacion,
                        "Gasolina de Salida": OPCIONES_GASOLINA[gas_salida_label],
                        "Gasolina de Llegada": OPCIONES_GASOLINA[gas_llegada_label],
                        "Dotación de Gasolina(LLENAR GASTO DE COMBUSTIBLE)": dotacion,
                        "Oficio Numero": oficio_num if oficio_num else None,
                        "Oficio Fecha": oficio_fecha if oficio_fecha else None,
                        "observaciones": observaciones if observaciones else None,
                        "Usuario Responsable": usuario,
                        "Áreas de Adscripción": residencia,
                        "Tipo de Vehículo": vehiculo,
                        "Placas": placas_upper,
                        "No. De Licencia": licencia,
                        "ESTADO_ADSCRIPCION": estado_usuario_actual,
                        "JEFATURA": jefatura_actual,
                        "JEFE_RESIDENCIA": jefe_actual,
                        "CORREO_ORGANIZADOR": current_email_key
                    }
                    
                    registros_actuales.append(nuevo_reg)
                    guardar_registros_acumulados(registros_actuales)
                    registrar_auditoria("CAPTURA RECORRIDO", f"Registro de {recorrido} km en {municipio} ({poblado})")
                    st.success(f"✅ ¡Día {fecha_str} en {municipio} ({poblado}) guardado correctamente!")

    registros_totales = cargar_registros_acumulados()
    if len(registros_totales) > 0:
        st.markdown("---")
        st.subheader("📋 Días Guardados (Historial Acumulado Permanente)")
        
        busqueda_texto = st.text_input("🔍 Buscar en historial (Folio CIIA, Oficio, Poblado o Municipio):").strip().lower()
        df_acumulado = pd.DataFrame(registros_totales)
        
        if busqueda_texto:
            mask = df_acumulado.apply(lambda row: row.astype(str).str.lower().str.contains(busqueda_texto).any(), axis=1)
            df_acumulado = df_acumulado[mask]
            
        st.dataframe(df_acumulado, use_container_width=True)
        
        col_acc1, col_acc2 = st.columns(2)
        with col_acc1:
            if st.button("🗑️ Limpiar historial completo"):
                guardar_registros_acumulados([])
                registrar_auditoria("LIMPIAR HISTORIAL", "Se eliminó todo el historial acumulado de recorridos")
                st.rerun()
        
        with col_acc2:
            if st.button("🚀 Guardar y Generar 3 Bitácoras Definitivas"):
                try:
                    wb = openpyxl.load_workbook("Prueba unificación.xlsx")
                    ws = wb["BASE_DE_DATOS"]
                    
                    # 1. Limpiar filas previas manteniendo intactas las demás hojas
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=25):
                        for cell in row:
                            cell.value = None
                            
                    # 2. Encabezados adicionales exigidos por las fórmulas de las hojas 2 y 3:
                    ws["K1"] = "Rendimiento (km/L)"
                    ws["L1"] = "Precio Gasolina ($)"
                    ws["M1"] = "GASTO COMBUSTI"
                    ws["N1"] = "Dotación de Gasolina(LLENAR GASTO DE COMBUSTIBLE)"
                    ws["O1"] = "Oficio Numero"
                    ws["P1"] = "Oficio Fecha"
                    ws["Q1"] = "observaciones"
                    ws["R1"] = "Usuario Responsable"
                    ws["S1"] = "Áreas de Adscripción"
                    ws["T1"] = "Tipo de Vehículo"
                    ws["U1"] = "Placas"
                    ws["V1"] = "No. De Licencia"
                    
                    # 3. Mapeo exacto de columnas acorde a las fórmulas de las hojas 2 y 3:
                    for i, reg in enumerate(registros_totales):
                        row_idx = 2 + i
                        
                        ws[f'A{row_idx}'] = reg["FECHA COMPLETA"]                                   # A: Fecha
                        ws[f'B{row_idx}'] = f'=UPPER(TEXT(A{row_idx}, "MMMM"))'                   # B: Mes
                        ws[f'C{row_idx}'] = reg["MUNICIPIO"]                                      # C: Municipio
                        ws[f'D{row_idx}'] = reg["POBLADO"]                                        # D: Poblado
                        ws[f'E{row_idx}'] = reg["folio CIIA"]                                     # E: Folio CIIA
                        ws[f'F{row_idx}'] = reg["HORA DE SALIDA"]                                 # F: Hora Salida
                        ws[f'G{row_idx}'] = reg["KM INICIAL / Km de Salida"]                      # G: Km Salida
                        ws[f'H{row_idx}'] = reg["RECORRIDO"]                                      # H: Recorrido (Km)
                        ws[f'I{row_idx}'] = reg["HORA DE LLEGADA"]                                # I: Hora Llegada
                        ws[f'J{row_idx}'] = reg["KM FINAL / Km de Llegada"]                       # J: Km Llegada
                        ws[f'K{row_idx}'] = 12.0                                                  # K: Rendimiento
                        ws[f'L{row_idx}'] = reg["Gasolina de Salida"]                             # L: Gas Salida
                        ws[f'M{row_idx}'] = f'=ROUND((H{row_idx}/12.0)*23.99, 2)'                 # M: Fórmula Gasto
                        ws[f'N{row_idx}'] = reg["Gasolina de Llegada"]                            # N: Gas Llegada
                        ws[f'O{row_idx}'] = reg["Oficio Numero"]                                  # O: Oficio Número
                        ws[f'P{row_idx}'] = reg["Oficio Fecha"]                                   # P: Oficio Fecha
                        ws[f'Q{row_idx}'] = reg["observaciones"]                                  # Q: Observaciones
                        ws[f'R{row_idx}'] = reg["Usuario Responsable"]                            # R: Usuario Responsable
                        ws[f'S{row_idx}'] = reg["Áreas de Adscripción"]                           # S: Áreas de Adscripción
                        ws[f'T{row_idx}'] = reg["Tipo de Vehículo"]                               # T: Tipo de Vehículo
                        ws[f'U{row_idx}'] = reg["Placas"]                                         # U: Placas
                        ws[f'V{row_idx}'] = reg["No. De Licencia"]                                # V: No. De Licencia

                    # 4. Guardar archivo y descargar
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    registrar_auditoria("GENERAR BITACORAS", "Generación exitosa con mapeo exacto de columnas")
                    st.success("¡Archivo generado con éxito y listo para descarga!")
                    st.download_button(
                        label="⬇️ Descargar Archivo Definitivo (Incluye las 3 Bitácoras)",
                        data=output,
                        file_name="BITACORAS_OFICIALES_DEFINITIVAS.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Error al generar el archivo definitivo: {e}")

elif perfil in ["Panel de Administración y Auditoría", "Panel de Supervisión (Estatal/Residencia)"]:
    st.subheader("📊 Panel de Gestión, Supervisión y Auditoría")
    
    if rol_actual == "Administrador Nacional":
        tab_reg_user, tab_edit_user, tab_ctrl_user, tab_resumen_auditoria, tab_bitacora_audit, tab_respaldo = st.tabs([
            "➕ Alta de Usuario", 
            "✏️ Editar Usuario", 
            "👥 Control y Estatus", 
            "📈 Resumen Ejecutivo y Auditoría",
            "🛡️ Bitácora de Auditoría",
            "💾 Respaldo de Usuarios"
        ])
    else:
        tab_resumen_auditoria = st.container()
        st.markdown("---")
        st.subheader("📋 Supervisión de Registros del Estado / Residencia")
        
    def render_alta_usuario():
        with st.form("form_nuevo_usuario"):
            col_u1, col_u2 = st.columns(2)
            with col_u1:
                c_email = st.text_input("Correo Electrónico (Usuario)", max_chars=300)
                c_nombre = st.text_input("Nombre Completo (Mayúsculas)", max_chars=300)
                c_pass = st.text_input("Contraseña Asignada", type="password", max_chars=300)
                c_licencia = st.text_input("Número de Licencia de Conducir", max_chars=300)
            with col_u2:
                c_estado = st.selectbox("Estado de Adscripción", ESTADOS_REPUBLICA)
                c_jefatura = st.selectbox("Jefatura de Residencia", JEFATURAS_RESIDENCIA)
                c_jefe_res = st.text_input("Nombre del Jefe de Residencia Asignado", value="ING. GABRIEL ESTRADA", max_chars=300)
                c_rol = st.selectbox("Rol en el Sistema", ROLES_SISTEMA)
            
            btn_crear = st.form_submit_button("Registrar Usuario en la Red")
            if btn_crear:
                if c_email and c_nombre and c_pass:
                    usuarios_actuales = cargar_usuarios()
                    email_limpio = c_email.strip().lower()
                    usuarios_actuales[email_limpio] = {
                        "nombre": c_nombre.strip().upper(),
                        "pass": c_pass.strip(),
                        "licencia": c_licencia.strip(),
                        "estado": c_estado,
                        "jefatura": c_jefatura,
                        "jefe_residencia": c_jefe_res.strip().upper(),
                        "rol": c_rol,
                        "foto": "",
                        "activo": True
                    }
                    guardar_usuarios(usuarios_actuales)
                    registrar_auditoria("ALTA USUARIO", f"Creación de usuario {email_limpio} con rol {c_rol}")
                    st.success(f"¡Usuario {c_nombre} ({c_rol}) registrado exitosamente para {c_estado}!")
                    st.rerun()
                else:
                    st.error("⚠️ Por favor completa los campos obligatorios (Correo, Nombre y Contraseña).")

    def render_editar_usuario():
        usuarios_actuales_edit = cargar_usuarios()
        lista_emails = list(usuarios_actuales_edit.keys())
        if lista_emails:
            email_a_editar = st.selectbox("Selecciona el correo del usuario a modificar", lista_emails)
            if email_a_editar:
                u_data = usuarios_actuales_edit[email_a_editar]
                with st.form("form_editar_usuario"):
                    col_e1, col_e2 = st.columns(2)
                    with col_e1:
                        e_nombre = st.text_input("Nombre Completo (Mayúsculas)", value=u_data.get("nombre", ""), max_chars=300)
                        e_pass = st.text_input("Contraseña Asignada", value=u_data.get("pass", ""), type="password", max_chars=300)
                        e_licencia = st.text_input("Número de Licencia de Conducir", value=u_data.get("licencia", ""), max_chars=300)
                    with col_e2:
                        estado_actual = u_data.get("estado", "Estado de México")
                        idx_estado = ESTADOS_REPUBLICA.index(estado_actual) if estado_actual in ESTADOS_REPUBLICA else 0
                        e_estado = st.selectbox("Estado de Adscripción", ESTADOS_REPUBLICA, index=idx_estado)
                        
                        jef_actual = u_data.get("jefatura", "RESIDENCIA NAUCALPAN")
                        idx_jef = JEFATURAS_RESIDENCIA.index(jef_actual) if jef_actual in JEFATURAS_RESIDENCIA else 0
                        e_jefatura = st.selectbox("Jefatura de Residencia", JEFATURAS_RESIDENCIA, index=idx_jef)
                        
                        e_jefe_res = st.text_input("Jefe de Residencia Asignado", value=u_data.get("jefe_residencia", ""), max_chars=300)
                        
                        idx_rol = ROLES_SISTEMA.index(u_data.get("rol", "Organizador Agrario (Operador)")) if u_data.get("rol") in ROLES_SISTEMA else 0
                        e_rol = st.selectbox("Rol en el Sistema", ROLES_SISTEMA, index=idx_rol)
                    
                    btn_actualizar = st.form_submit_button("💾 Guardar Cambios")
                    if btn_actualizar:
                        usuarios_actuales_edit[email_a_editar]["nombre"] = e_nombre.strip().upper()
                        usuarios_actuales_edit[email_a_editar]["pass"] = e_pass.strip()
                        usuarios_actuales_edit[email_a_editar]["licencia"] = e_licencia.strip()
                        usuarios_actuales_edit[email_a_editar]["estado"] = e_estado
                        usuarios_actuales_edit[email_a_editar]["jefatura"] = e_jefatura
                        usuarios_actuales_edit[email_a_editar]["jefe_residencia"] = e_jefe_res.strip().upper()
                        usuarios_actuales_edit[email_a_editar]["rol"] = e_rol
                        guardar_usuarios(usuarios_actuales_edit)
                        registrar_auditoria("EDITAR USUARIO", f"Actualización de datos para {email_a_editar}")
                        st.success(f"¡Información de {email_a_editar} actualizada exitosamente!")
                        st.rerun()
        else:
            st.info("No hay usuarios registrados para editar.")

    def render_control_estatus():
        st.subheader("👥 Control, Estatus y Eliminación de Usuarios en la Red Nacional")
        usuarios_actuales_tabla = cargar_usuarios()
        
        filas_html = ""
        for email, datos in usuarios_actuales_tabla.items():
            estado_activo = datos.get("activo", True)
            color_fondo = "#d4edda" if estado_activo else "#e2e3e5"
            texto_estado = "🟢 Activo" if estado_activo else "🔴 Desac."
            
            filas_html += f"""
            <div style="background-color: {color_fondo}; padding: 8px 12px; border-radius: 6px; margin-bottom: 6px; border: 1px solid #c0c0c0; display: flex; justify-content: space-between; align-items: center; font-size: 12px;">
                <span style="flex: 2.0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; text-align: left;"><b>{email}</b></span>
                <span style="flex: 2.5; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; text-align: left;">{datos.get('nombre')}</span>
                <span style="flex: 2.0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; text-align: left;">{datos.get('rol')}</span>
                <span style="flex: 1.3; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; text-align: left;">{datos.get('estado')}</span>
                <span style="flex: 1.8; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; text-align: left;">{datos.get('jefatura')}</span>
                <span style="flex: 0.9; text-align: right; white-space: nowrap;"><b>{texto_estado}</b></span>
            </div>
            """
            
        st.markdown(
            f"""
            <div style="background-color: #6B1D2F; color: white; padding: 10px 15px; border-radius: 6px; font-weight: bold; display: flex; justify-content: space-between; align-items: center; margin-bottom: 6px; font-size: 13px;">
                <span style="flex: 2.0; text-align: left;">CORREO</span>
                <span style="flex: 2.5; text-align: left;">NOMBRE</span>
                <span style="flex: 2.0; text-align: left;">ROL</span>
                <span style="flex: 1.3; text-align: left;">ESTADO</span>
                <span style="flex: 1.8; text-align: left;">JEFATURA</span>
                <span style="flex: 0.9; text-align: right;">ESTATUS</span>
            </div>
            {filas_html}
            """,
            unsafe_allow_html=True
        )
        
        st.markdown("---")
        st.subheader("⚙️ Botones de Acción por Usuario")
        for email, datos in usuarios_actuales_tabla.items():
            col_acc_lbl, col_b1, col_b2, col_b3 = st.columns([3, 1, 1, 1])
            with col_acc_lbl:
                st.write(f"**{email}** ({datos.get('nombre')})")
            with col_b1:
                if st.button("🟢 Encender", key=f"activar_{email}"):
                    usuarios_actuales_tabla[email]["activo"] = True
                    guardar_usuarios(usuarios_actuales_tabla)
                    registrar_auditoria("ACTIVAR USUARIO", f"Activación de cuenta para {email}")
                    st.rerun()
            with col_b2:
                if st.button("🔴 Apagar", key=f"desactivar_{email}"):
                    usuarios_actuales_tabla[email]["activo"] = False
                    guardar_usuarios(usuarios_actuales_tabla)
                    registrar_auditoria("DESACTIVAR USUARIO", f"Desactivación de cuenta para {email}")
                    st.rerun()
            with col_b3:
                if st.button("🗑️ Eliminar", key=f"eliminar_{email}"):
                    if email == st.session_state["current_email"]:
                        st.error("No puedes eliminar tu propia cuenta.")
                    else:
                        del usuarios_actuales_tabla[email]
                        guardar_usuarios(usuarios_actuales_tabla)
                        registrar_auditoria("ELIMINAR USUARIO", f"Eliminación permanente de cuenta para {email}")
                        st.rerun()

    def render_respaldo_usuarios():
        st.subheader("💾 Respaldo y Restauración de Usuarios")
        st.markdown("Descarga la base de datos actual de usuarios o sube un archivo previamente guardado para restaurarlos.")
        
        usuarios_actuales = cargar_usuarios()
        json_usuarios = json.dumps(usuarios_actuales, ensure_ascii=False, indent=4)
        
        st.download_button(
            label="⬇️ Descargar Archivo de Usuarios (usuarios_respaldo.json)",
            data=json_usuarios,
            file_name="usuarios_respaldo.json",
            mime="application/json"
        )
        
        st.markdown("---")
        st.write("**Restaurar Usuarios desde Archivo**")
        archivo_respaldo = st.file_uploader("Sube el archivo usuarios_respaldo.json que descargaste previamente", type=["json"])
        
        if st.button("📤 Cargar y Restaurar Usuarios"):
            if archivo_respaldo is not None:
                try:
                    usuarios_cargados = json.load(archivo_respaldo)
                    guardar_usuarios(usuarios_cargados)
                    registrar_auditoria("RESTAURAR USUARIOS", "Se restauró la base de datos de usuarios desde un archivo externo")
                    st.success("✅ ¡Usuarios restaurados con éxito! El sistema aplicará los cambios ahora.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al leer el archivo. Asegúrate de que es un JSON válido. Detalle: {e}")
            else:
                st.warning("⚠️ Selecciona un archivo de tu computadora primero.")

    def render_resumen_auditoria():
        st.subheader("📈 Resumen Ejecutivo y Auditoría (Control Vehicular)")
        st.markdown("Consulta y segmentación del uso general de vehículos por rango de fechas, usuario organizador, estado y jefatura de residencia.")
        
        registros_persisted = cargar_registros_acumulados()
        if len(registros_persisted) > 0:
            df_global = pd.DataFrame(registros_persisted)
            
            for col in ['JEFATURA', 'ESTADO_ADSCRIPCION', 'JEFE_RESIDENCIA', 'Áreas de Adscripción', 'Usuario Responsable']:
                if col not in df_global.columns:
                    df_global[col] = 'N/A'
            
            df_global['FECHA_DT'] = pd.to_datetime(df_global['FECHA COMPLETA'], format='%d/%m/%Y', errors='coerce').dt.date
            
            if rol_actual in ["Jefe de Residencia", "Administrador de Vehículos (Estatal)"]:
                jefatura_sesion = st.session_state.get("current_jefatura", "")
                df_global = df_global[df_global['JEFATURA'] == jefatura_sesion]
                st.info(f"Mostrando registros bajo la Jefatura: **{jefatura_sesion}**")
            elif rol_actual in ["Administrador Estatal", "Analista de Información"]:
                estado_sesion = st.session_state.get("current_estado", "")
                df_global = df_global[df_global['ESTADO_ADSCRIPCION'] == estado_sesion]
                st.info(f"Mostrando registros del Estado: **{estado_sesion}**")
            
            col_f1, col_f2, col_f3, col_f4, col_f5 = st.columns(5)
            with col_f1:
                f_inicio = st.date_input("Fecha Inicial", value=df_global['FECHA_DT'].min() if not df_global.empty else date.today())
            with col_f2:
                f_fin = st.date_input("Fecha Final", value=df_global['FECHA_DT'].max() if not df_global.empty else date.today())
            with col_f3:
                orgs_disp = ["Todos los organizadores"] + sorted(df_global['Usuario Responsable'].dropna().unique().tolist()) if not df_global.empty else ["Todos"]
                filtro_usuario = st.selectbox("Organizador Agrario", orgs_disp)
            with col_f4:
                est_disp = ["Todos los estados"] + sorted(df_global['ESTADO_ADSCRIPCION'].dropna().unique().tolist()) if not df_global.empty else ["Todos"]
                filtro_estado = st.selectbox("Estado", est_disp)
            with col_f5:
                jef_disp = ["Todas las jefaturas"] + sorted(df_global['JEFATURA'].dropna().unique().tolist()) if not df_global.empty else ["Todas"]
                filtro_jefatura = st.selectbox("Jefatura de Residencia", jef_disp)
                
            if not df_global.empty:
                df_filtrado = df_global[(df_global['FECHA_DT'] >= f_inicio) & (df_global['FECHA_DT'] <= f_fin)]
                if filtro_usuario != "Todos los organizadores":
                    df_filtrado = df_filtrado[df_filtrado['Usuario Responsable'] == filtro_usuario]
                if filtro_estado != "Todos los estados":
                    df_filtrado = df_filtrado[df_filtrado['ESTADO_ADSCRIPCION'] == filtro_estado]
                if filtro_jefatura != "Todas las jefaturas":
                    df_filtrado = df_filtrado[df_filtrado['JEFATURA'] == filtro_jefatura]
                
                if not df_filtrado.empty:
                    st.markdown("---")
                    total_km = df_filtrado['RECORRIDO'].sum()
                    total_gas = df_filtrado['GASTO COMBUSTI'].sum()
                    total_viajes = len(df_filtrado)
                    
                    m1, m2, m3 = st.columns(3)
                    m1.metric("KM Totales Recorridos", f"{total_km:,.1f} km")
                    m2.metric("Gasto Total de Combustible", f"${total_gas:,.2f} MXN")
                    m3.metric("Comisiones Registradas", f"{total_viajes}")
                    
                    st.markdown("---")
                    st.subheader("🔧 Alertas de Mantenimiento Preventivo (Flotilla)")
                    df_mantenimiento = df_filtrado.groupby("Placas")["RECORRIDO"].sum().reset_index()
                    for _, row in df_mantenimiento.iterrows():
                        km_total_veh = row["RECORRIDO"]
                        placa_v = row["Placas"]
                        if km_total_veh >= 5000:
                            st.error(f"🚨 **ALERTA DE MANTENIMIENTO**: El vehículo con placas **{placa_v}** ha acumulado **{km_total_veh:,.1f} km**. Requiere servicio preventivo urgente (cambio de aceite/filtros).")
                        else:
                            st.info(f"✅ El vehículo con placas **{placa_v}** registra **{km_total_veh:,.1f} km** (En rango operativo normal).")
                    
                    st.markdown("---")
                    st.subheader("📊 Gráficos Analíticos de Kilometraje, Gasto y Tendencia")
                    
                    col_g1, col_g2 = st.columns(2)
                    with col_g1:
                        st.write("**Kilómetros por Municipio**")
                        df_mun_km = df_filtrado.groupby("MUNICIPIO")["RECORRIDO"].sum()
                        st.bar_chart(df_mun_km)
                    with col_g2:
                        st.write("**Gasto de Combustible por Jefatura**")
                        df_jef_gas = df_filtrado.groupby("JEFATURA")["GASTO COMBUSTI"].sum()
                        st.bar_chart(df_jef_gas)
                    
                    st.write("**Evolución Temporal de Kilómetros Recorridos**")
                    df_temporal = df_filtrado.groupby("FECHA COMPLETA")["RECORRIDO"].sum()
                    st.line_chart(df_temporal)
                    
                    st.markdown("---")
                    st.subheader("🗺️ Mapa Interactivo de Movimiento y Rutas (Recorridos)")
                    st.markdown("Visualización cartográfica de los destinos visitados y líneas de conexión cronológica entre comisiones.")
                    
                    mapa_rows = []
                    for _, r in df_filtrado.iterrows():
                        mun = r["MUNICIPIO"]
                        if mun in CODS_MUNICIPIOS:
                            mapa_rows.append({
                                "lat": CODS_MUNICIPIOS[mun]["lat"],
                                "lon": CODS_MUNICIPIOS[mun]["lon"],
                                "MUNICIPIO": mun,
                                "POBLADO": r["POBLADO"],
                                "FECHA": r["FECHA COMPLETA"],
                                "USUARIO": r["Usuario Responsable"],
                                "RECORRIDO": r["RECORRIDO"]
                            })
                    
                    if mapa_rows:
                        df_mapa = pd.DataFrame(mapa_rows)
                        arcos_rows = []
                        if len(df_mapa) > 1:
                            for i in range(len(df_mapa) - 1):
                                arcos_rows.append({
                                    "start_lon": df_mapa.iloc[i]["lon"],
                                    "start_lat": df_mapa.iloc[i]["lat"],
                                    "end_lon": df_mapa.iloc[i+1]["lon"],
                                    "end_lat": df_mapa.iloc[i+1]["lat"]
                                })
                        df_arcos = pd.DataFrame(arcos_rows) if arcos_rows else pd.DataFrame()
                        
                        capa_puntos = pdk.Layer(
                            "ScatterplotLayer",
                            data=df_mapa,
                            get_position='[lon, lat]',
                            get_color='[107, 29, 47, 180]',
                            get_radius=3000,
                            pickable=True,
                            auto_highlight=True
                        )
                        
                        capa_lineas = None
                        if not df_arcos.empty:
                            capa_lineas = pdk.Layer(
                                "ArcLayer",
                                data=df_arcos,
                                get_source_position='[start_lon, start_lat]',
                                get_target_position='[end_lon, end_lat]',
                                get_source_color='[107, 29, 47]',
                                get_target_color='[138, 37, 61]',
                                get_width=3
                            )
                        
                        lat_centro = df_mapa["lat"].mean()
                        lon_centro = df_mapa["lon"].mean()
                        
                        view_state = pdk.ViewState(
                            latitude=lat_centro,
                            longitude=lon_centro,
                            zoom=8,
                            pitch=30
                        )
                        
                        layers = [capa_puntos]
                        if capa_lineas:
                            layers.append(capa_lineas)
                            
                        r_deck = pdk.Deck(
                            layers=layers,
                            initial_view_state=view_state,
                            tooltip={
                                "html": "<b>Municipio:</b> {MUNICIPIO}<br/><b>Poblado:</b> {POBLADO}<br/><b>Fecha:</b> {FECHA}<br/><b>Responsable:</b> {USUARIO}",
                                "style": {"backgroundColor": "#6B1D2F", "color": "white"}
                            }
                        )
                        st.pydeck_chart(r_deck)
                    else:
                        st.info("No hay coordenadas registradas para los municipios seleccionados en el filtro.")
                    
                    st.markdown("---")
                    st.subheader("📋 Trazabilidad Geográfica y Municipios Visitados")
                    
                    df_resumen_mun = df_filtrado.groupby(['FECHA COMPLETA', 'MUNICIPIO', 'POBLADO', 'Usuario Responsable', 'JEFATURA', 'ESTADO_ADSCRIPCION']).agg({
                        'RECORRIDO': 'sum',
                        'GASTO COMBUSTI': 'sum'
                    }).reset_index().sort_values(by='FECHA COMPLETA')
                    
                    st.dataframe(df_resumen_mun, use_container_width=True)
                    
                    municipios_visitados = df_filtrado['MUNICIPIO'].unique().tolist()
                    st.markdown(f"**Municipios únicos visitados en el filtro ({len(municipios_visitados)}):** " + ", ".join(municipios_visitados))
                    
                    st.markdown("---")
                    output_filtrado = BytesIO()
                    df_resumen_mun.to_excel(output_filtrado, index=False, sheet_name="REPORTE_FILTRADO")
                    output_filtrado.seek(0)
                    st.download_button(
                        label="⬇️ Descargar Reporte Filtrado en Excel",
                        data=output_filtrado,
                        file_name="REPORTE_EJECUTIVO_FILTRADO.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("⚠️ No se encontraron registros con los filtros seleccionados.")
            else:
                st.warning("⚠️ No hay registros disponibles para tu nivel de adscripción.")
        else:
            st.info("Aún no hay registros de recorridos acumulados en el sistema.")

    def render_bitacora_audit():
        st.subheader("🛡️ Bitácora de Auditoría del Sistema")
        st.markdown("Registro cronológico de acciones administrativas y de seguridad realizadas en la red nacional.")
        if os.path.exists(AUDIT_FILE):
            try:
                with open(AUDIT_FILE, "r", encoding="utf-8") as f:
                    logs_data = json.load(f)
                if logs_data:
                    df_audit = pd.DataFrame(logs_data)
                    st.dataframe(df_audit, use_container_width=True)
                else:
                    st.info("No hay registros de auditoría almacenados.")
            except:
                st.info("No se pudo leer el archivo de auditoría.")
        else:
            st.info("Aún no se han generado eventos de auditoría.")

    if rol_actual == "Administrador Nacional":
        with tab_reg_user:
            render_alta_usuario()
        with tab_edit_user:
            render_editar_usuario()
        with tab_ctrl_user:
            render_control_estatus()
        with tab_resumen_auditoria:
            render_resumen_auditoria()
        with tab_bitacora_audit:
            render_bitacora_audit()
        with tab_respaldo:
            render_respaldo_usuarios()
    else:
        render_resumen_auditoria()
